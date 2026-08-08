"""Parser for a personal finance-tracking spreadsheet exported as .xlsx.

This is a snapshot-upload integration, not a live Google Sheets API
connection — see README.md > "Personal finance sheet" for the refresh
workflow (download the sheet as .xlsx from Google Drive, upload it here,
same pattern as a bank statement).

Two tabs are deliberately never opened: anything matching SKIP_SHEET_PATTERNS
(trading / savings-scheme tracking) is skipped entirely, per an explicit
user request to exclude them from this tool.

The layout below (section headers, column positions) was reverse-engineered
from one real spreadsheet and is inherently specific to it — unlike the bank
statement parsers, there's no "generic personal tracker" format to target.
If your sheet's tab names or layout differ, treat this as a starting point:
each `parse_*` function documents the exact layout it expects.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import pandas as pd

SKIP_SHEET_PATTERNS = [r"trading", r"saving.?s?_?scheme"]

# --- Investments tab -------------------------------------------------------

SECTION_HEADERS = {
    "Mutual Funds", "Fixed Deposits", "Recurring Deposits", "Savings",
    "Others", "Provident Fund", "Other Income", "Fixed Assets", "Liabilities",
}
LIABILITY_SUBSECTIONS = {"Loans", "Credit Cards"}


@dataclass
class InvestmentItem:
    section: str
    subsection: Optional[str]
    name: str
    amount: Optional[float]        # column B — principal / monthly SIP / one-off amount
    current_value: Optional[float]  # column C — current market value, where tracked
    monthly_value: Optional[float]  # column D — recurring monthly income (rent etc.)
    details: Optional[str]          # column F — folio numbers / free-text notes


def _num(v):
    return v if isinstance(v, (int, float)) else None


def parse_investments_tab(ws) -> list[InvestmentItem]:
    """Expects: col A=item/section name, B=Amount, C=Current Value,
    D=Monthly Values, F=Details. Section headers and sub-section markers
    (leading '-', or 'Loans'/'Credit Cards' under Liabilities) switch
    context; rows with no label (blank separators or the sheet's own
    hand-computed subtotals) are skipped — this module derives totals by
    summing the parsed items instead of scraping those subtotal cells, which
    is more robust to the sheet's inconsistent formatting.
    """
    items: list[InvestmentItem] = []
    section: Optional[str] = None
    subsection: Optional[str] = None

    for row in ws.iter_rows(min_row=2):
        a = row[0].value if len(row) > 0 else None
        b = row[1].value if len(row) > 1 else None
        c = row[2].value if len(row) > 2 else None
        d = row[3].value if len(row) > 3 else None
        f = row[5].value if len(row) > 5 else None

        label = a.strip() if isinstance(a, str) else None
        if not label:
            continue

        if label in SECTION_HEADERS:
            section, subsection = label, None
            continue
        if label in LIABILITY_SUBSECTIONS:
            subsection = label
            continue
        if label.startswith("-"):
            subsection = label.lstrip("- ").strip()
            continue
        if label.upper().startswith("TOTAL"):
            continue

        items.append(
            InvestmentItem(
                section=section or "Uncategorized",
                subsection=subsection,
                name=label,
                amount=_num(b),
                current_value=_num(c),
                monthly_value=_num(d),
                details=str(f).strip() if f not in (None, "") else None,
            )
        )
    return items


def investment_items_to_frame(items: list[InvestmentItem]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "section": i.section,
            "subsection": i.subsection,
            "name": i.name,
            "amount": i.amount,
            "current_value": i.current_value,
            "monthly_value": i.monthly_value,
            "details": i.details,
        }
        for i in items
    )


def section_totals(items: list[InvestmentItem]) -> pd.DataFrame:
    """Sum by section, preferring current_value where the sheet tracks it
    (mutual funds, PF) and falling back to amount (FDs, savings, real
    estate, loans — where the sheet only records a principal)."""
    rows = []
    for section in sorted({i.section for i in items}):
        sub = [i for i in items if i.section == section]
        value = sum((i.current_value or i.amount or 0) for i in sub)
        rows.append({"section": section, "total": value, "count": len(sub)})
    return pd.DataFrame(rows).sort_values("total", ascending=False)


# --- MonthlyYearly Expenses tab ---------------------------------------------

_BLOCK3_SKIP = {"details", "ott"}  # "OTT" here is a subtotal row, not a line item


def parse_budget_tab(ws) -> pd.DataFrame:
    """Expects three side-by-side blocks on one sheet:
    - cols A/B(/C): recurring monthly allocations (SIPs, EMIs, house fund),
      with a second sub-block further down giving yearly totals directly.
    - cols D/E: income sources.
    - cols H/I/J: itemized house expenses and subscriptions (monthly +
      yearly columns).
    Blank-labelled rows are the sheet's own subtotals and are skipped in
    favour of summing the parsed rows.
    """
    records = []

    for row in ws.iter_rows(min_row=2, max_row=20):
        name = row[0].value if len(row) > 0 else None
        monthly = row[1].value if len(row) > 1 else None
        yearly = row[2].value if len(row) > 2 else None
        if isinstance(name, str) and name.strip() and _num(monthly) is not None:
            m = _num(monthly)
            y = _num(yearly) if _num(yearly) is not None else round(m * 12, 2)
            records.append({"category": name.strip(), "type": "Allocation", "monthly_amount": m, "yearly_amount": y})

    for row in ws.iter_rows(min_row=2, max_row=20):
        name = row[3].value if len(row) > 3 else None
        amount = row[4].value if len(row) > 4 else None
        if isinstance(name, str) and name.strip() and _num(amount) is not None:
            m = _num(amount)
            records.append({"category": name.strip(), "type": "Income", "monthly_amount": m, "yearly_amount": round(m * 12, 2)})

    for row in ws.iter_rows(min_row=2, max_row=20):
        name = row[7].value if len(row) > 7 else None
        monthly = row[8].value if len(row) > 8 else None
        yearly = row[9].value if len(row) > 9 else None
        if not isinstance(name, str) or not name.strip():
            continue
        if name.strip().lower() in _BLOCK3_SKIP:
            continue
        m, y = _num(monthly), _num(yearly)
        if m is None and y is None:
            continue
        if m is None:
            m = round(y / 12, 2)
        if y is None:
            y = round(m * 12, 2)
        records.append({"category": name.strip(), "type": "House Expense", "monthly_amount": m, "yearly_amount": y})

    return pd.DataFrame(records)


# --- NirmanPayments tab -----------------------------------------------------

_SUMMARY_LABELS = {"Loan Disbursements", "Loan Principal", "Self Payments", "Loan Pending", "Total", "Pending"}


def parse_nirman_tab(ws) -> tuple[pd.DataFrame, dict[str, float]]:
    """Expects an EMI amortization schedule in columns H (date), I (EMI
    total, sparsely filled), J (interest), K (principal) — columns 7-10,
    0-indexed — and a "Summary" block in columns A/B(/C) further down with
    fixed labels (Loan Disbursements, Loan Principal, Self Payments, Loan
    Pending, Total, Pending). Everything else on this sheet (TDS breakdown,
    disbursement timeline, ad-hoc payment notes) is too unstructured to
    parse reliably and is left out — see the sheet directly for that detail.
    """
    schedule_rows = []
    summary: dict[str, float] = {}

    for row in ws.iter_rows(min_row=2):
        date_val = row[7].value if len(row) > 7 else None
        total_val = _num(row[8].value) if len(row) > 8 else None
        interest = _num(row[9].value) if len(row) > 9 else None
        principal = _num(row[10].value) if len(row) > 10 else None
        if date_val is not None and (interest is not None or principal is not None):
            emi_total = total_val if total_val is not None else round((interest or 0) + (principal or 0), 2)
            schedule_rows.append({"date": date_val, "emi_total": emi_total, "interest": interest or 0, "principal": principal or 0})

        label = row[0].value if len(row) > 0 else None
        if isinstance(label, str) and label.strip() in _SUMMARY_LABELS:
            b, c = row[1].value if len(row) > 1 else None, row[2].value if len(row) > 2 else None
            val = _num(b) if _num(b) is not None else _num(c)
            if val is not None:
                summary[label.strip()] = val

    schedule = pd.DataFrame(schedule_rows)
    if not schedule.empty:
        schedule["date"] = pd.to_datetime(schedule["date"])
        schedule = schedule.sort_values("date").reset_index(drop=True)
    return schedule, summary


# --- My_Savings_Investments tab --------------------------------------------

def parse_savings_summary_tab(ws) -> dict:
    """Pulls just the pieces of this tab that aren't already covered by the
    Investments tab: the hand-computed "Total Surplus" figure, and the
    "Devika" (partner) equity-share breakdown by property. The rest of this
    tab is a re-aggregation of the Investments tab and is skipped to avoid
    presenting the same numbers twice under different labels.
    """
    total_surplus = None
    devika_equity: dict[str, float] = {}
    in_devika_block = False

    for row in ws.iter_rows(min_row=1):
        vals = [c.value for c in row]
        for i, v in enumerate(vals):
            if v == "Total Surplus":
                for later in vals[i + 1:]:
                    if _num(later) is not None:
                        total_surplus = _num(later)
                        break

        label = vals[0].strip() if isinstance(vals[0], str) else None
        if label == "Devika":
            in_devika_block = True
            continue
        if in_devika_block and label:
            amt = _num(vals[1]) if len(vals) > 1 else None
            if amt is not None:
                devika_equity[label] = amt

    return {"total_surplus": total_surplus, "devika_equity": devika_equity}


# --- top-level entry point --------------------------------------------------

@dataclass
class PersonalFinanceData:
    investments: list[InvestmentItem] = field(default_factory=list)
    budget: pd.DataFrame = field(default_factory=pd.DataFrame)
    nirman_schedule: pd.DataFrame = field(default_factory=pd.DataFrame)
    nirman_summary: dict = field(default_factory=dict)
    savings_summary: dict = field(default_factory=dict)
    skipped_sheets: list[str] = field(default_factory=list)
    sheets_found: list[str] = field(default_factory=list)


def load_personal_sheet(file_bytes: bytes) -> PersonalFinanceData:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    data = PersonalFinanceData(sheets_found=list(wb.sheetnames))

    for name in wb.sheetnames:
        if any(re.search(p, name.lower()) for p in SKIP_SHEET_PATTERNS):
            data.skipped_sheets.append(name)
            continue

        ws = wb[name]
        key = name.lower().replace(" ", "").replace("_", "")
        if key == "investments":
            data.investments = parse_investments_tab(ws)
        elif "monthlyyearlyexpenses" in key or key == "expenses":
            data.budget = parse_budget_tab(ws)
        elif "nirman" in key:
            data.nirman_schedule, data.nirman_summary = parse_nirman_tab(ws)
        elif "savingsinvestments" in key or "mysavings" in key:
            data.savings_summary = parse_savings_summary_tab(ws)

    return data
